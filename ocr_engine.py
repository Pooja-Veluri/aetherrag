import io
import logging
from PIL import Image
import pdfplumber
import pytesseract
from pdf2image import convert_from_bytes
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

def extract_text_from_pdf_native(pdf_bytes: bytes) -> list[dict]:
    """
    Extracts text from a PDF file using pdfplumber (native text extraction).
    Returns a list of dicts: [{"page": page_num, "text": text}, ...]
    """
    pages_data = []
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if text:
                    text = text.strip()
                else:
                    text = ""
                pages_data.append({"page": i + 1, "text": text})
    except Exception as e:
        logger.error(f"Error during native PDF extraction: {e}")
        raise e
    return pages_data

def extract_text_from_pdf_ocr(pdf_bytes: bytes) -> list[dict]:
    """
    Converts PDF pages to images and runs Tesseract OCR on each page.
    Returns a list of dicts: [{"page": page_num, "text": text}, ...]
    """
    pages_data = []
    try:
        # Convert PDF bytes to PIL Images
        images = convert_from_bytes(pdf_bytes)
        for i, img in enumerate(images):
            # Run pytesseract OCR on each image
            text = pytesseract.image_to_string(img)
            pages_data.append({
                "page": i + 1,
                "text": text.strip() if text else ""
            })
    except Exception as e:
        logger.error(f"Error during PDF OCR extraction: {e}")
        raise e
    return pages_data

def extract_text_from_pdf(pdf_bytes: bytes) -> list[dict]:
    """
    Tries native text extraction first. If the total extracted text is empty or
    extremely short, falls back to PDF OCR.
    """
    pages_data = extract_text_from_pdf_native(pdf_bytes)
    total_length = sum(len(page["text"]) for page in pages_data)
    
    # If the text is empty or very short (e.g. image-only PDF), run OCR
    if total_length < 100:
        logger.info("Native text extraction returned very little text. Falling back to OCR.")
        try:
            pages_data = extract_text_from_pdf_ocr(pdf_bytes)
        except Exception as ocr_err:
            logger.warning(f"PDF OCR failed: {ocr_err}. Returning native results.")
            
    return pages_data

def extract_text_from_image(image_bytes: bytes) -> list[dict]:
    """
    Extracts text from an image (PNG, JPEG, etc.) using Tesseract OCR.
    Returns a single-page list structure: [{"page": 1, "text": text}]
    """
    try:
        image = Image.open(io.BytesIO(image_bytes))
        text = pytesseract.image_to_string(image)
        return [{"page": 1, "text": text.strip() if text else ""}]
    except Exception as e:
        logger.error(f"Error during image OCR extraction: {e}")
        raise e

def extract_text_from_xlsx(xlsx_bytes: bytes) -> list[dict]:
    """
    Extract text from an Excel workbook.

    Emits one entry per data row (not per sheet) so the chunker has natural
    row-level boundaries to pack into chunks. Sheets without detectable headers
    fall back to pipe-joined cells. The sheet name is preserved by prefixing
    each row's text with a markdown section header so the chunker's section
    detector picks it up.

    Each emitted row is a "page" in the pipeline's vocabulary; in practice
    several rows will be packed into one retrieval chunk by the chunker.
    """
    def _is_numeric(v: str) -> bool:
        try:
            float(v)
            return True
        except ValueError:
            return False

    pages_data = []
    page_counter = 0
    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [
                [("" if c is None else str(c).strip()) for c in row]
                for row in ws.iter_rows(values_only=True)
            ]
            rows = [r for r in rows if any(cell for cell in r)]
            if not rows:
                continue

            first = rows[0]
            looks_like_header = (
                len(first) > 0
                and all(cell for cell in first)
                and len(set(first)) == len(first)
                and not any(_is_numeric(cell) for cell in first)
            )
            headers = first if looks_like_header else None
            body_rows = rows[1:] if looks_like_header else rows

            for row in body_rows:
                if headers:
                    pairs = [f"{h}: {v}" for h, v in zip(headers, row) if v]
                    if not pairs:
                        continue
                    row_text = " | ".join(pairs)
                else:
                    cells = [v for v in row if v]
                    if not cells:
                        continue
                    row_text = " | ".join(cells)

                page_counter += 1
                # Prefix the sheet name as a markdown header on the first row of
                # each sheet so the chunker tags subsequent chunks with `section=sheet_name`.
                # Subsequent rows just carry their data; the chunker's last-seen
                # section is sticky.
                if not pages_data or pages_data[-1].get("_sheet") != sheet_name:
                    text = f"## {sheet_name}\n{row_text}"
                else:
                    text = row_text
                pages_data.append({"page": page_counter, "text": text, "_sheet": sheet_name})
        wb.close()
    except Exception as e:
        logger.error(f"Error extracting xlsx: {e}")
        raise

    # Strip the internal _sheet marker before returning
    for p in pages_data:
        p.pop("_sheet", None)
    return pages_data


def extract_text(file_bytes: bytes, file_name: str) -> list[dict]:
    """
    Generic routing function to extract text from files by name/type.
    """
    ext = file_name.split(".")[-1].lower()
    if ext == "pdf":
        return extract_text_from_pdf(file_bytes)
    elif ext in ["png", "jpg", "jpeg", "webp", "tiff", "bmp"]:
        return extract_text_from_image(file_bytes)
    elif ext in ["xlsx", "xlsm"]:
        return extract_text_from_xlsx(file_bytes)
    else:
        # Fallback for plain text files just in case
        try:
            text = file_bytes.decode("utf-8")
            return [{"page": 1, "text": text}]
        except Exception as e:
            raise ValueError(f"Unsupported file format or encoding: {file_name}. Error: {e}")
